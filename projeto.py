import pyautogui as pug
# biblioteca .py - interface gráfica - controle de mouse, teclado e interação com atela
import time as tm
# biblioteca de manipulação de tempo, ex sleep
import os
# função que interage com sp - gerenciamento de files como criar, deletar e seus caminhos
import pandas as pd
# biblioteca para nanipular dados
from openpyxl import load_workbook
# carrega planilha excel .xlsx
# worklbook = tabela
from openpyxl.styles import Font, PatternFill
# Font fontes e PatternFill cores de fundos
from openpyxl.utils import get_column_letter
# tem haver com colunas

# ___________________________________ CRIANDO XLSX ___________________________________________________________________#

tabela = {
    "tarefa":["Abrir file", "salvar file", "deletar file", "fechar file", "editar file", "copiar file", "colar file", "nomear file"],
    "tipo":["click", "hotkey", "click", "click", "click", "hotkey", "hotkey", "textor" ],
    "dados":["tabela.xlsx", "ctrl+s", "save", "delete", "close", "ctrl+c", "ctrl+v", "nome"]
}
#cuidado com pontuação errada : iníco na chave e , entre dados


# _____________________________________________ CRIANDO DF/DATAFRAME ____________________________________________________________#

df = pd.DataFrame(tabela)
print(df)                              


#dentro do objeto dataframe passar as chaves/ tarefa, tipo e dados, conteúdo da coluna é o que está dentro das chaves


# _____________________________________________ SALVANDDO EM EXCEL _______________________________________________________________#

arq_excel = "tabela.xlsx"

df.to_excel(arq_excel, index=False, sheet_name="tabela_RPA")

# df.to_excel = transformando em arq excel | sheet_name = 1° planilha se chamará tabelaRPA 
# arq_excel é minga tabela.xlsx | index= False = não tem chave primária - se tiver deve por True

# __________________________________________________ FORMATAR ARQ _________________________________________________________________#

wb = load_workbook(arq_excel)
ws = wb["tabela_RPA"]

# wb = worldworkbook
# load_workbook = carrega o arq excel
# ws = worksheet - aba específica - lembrando que ao criar vem com 3 abas


# header_fill = PatternFill(start_color="edede9", end_color="f5ebe0", fill_type="gradient") # fundo de gradiente
#cor row = linha completa
header_font = Font(color="253237",bold=True)
header_fill = PatternFill(start_color="d8e4e8", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
# configuração e aplicação - cabeçalho
# usa-se o for cel porque deverá, na 1° linha, percorrer todas suas células e aplicar, "for" horizontal
#cor row = linha completa

#linha zebrada com sim cor não
lin_par_c = PatternFill(start_color="e0fbfc", fill_type="solid")
lin_inp_c = PatternFill(start_color="c2dfe3", fill_type="solid") 
for row in ws.iter_rows(min_row=1): 
    for cell in row:
        if cell.row % 2 == 0:
            cell.fill = lin_par_c
        else:
            cell.fill = lin_inp_c



     

# header_fill/ define cor de fundo - star color = inicia com esta cor   | fill_type =   | end_color =
# header_fonte =
# for cell in ws/worksheet = 
# _____________________________________________________ POSIÇÃO DO MOUSE ___________________________________________________________#  

print("Mostrando posição do mouse")
print("Mova o cursor do maouse")

try:
    while True:
        x, y = pug.position()
        # biblioteca pyautogui = pug  
        # print(f"\nLocal atual x = {x} e y = {y}", end="\r") -> aqui dá uma lista de posições, quero apenas posição atual
        print(f"Local atual x = {x} e y = {y}", end="\r")
        #aqui muda a posição sem mostrar o "histórico" de posição
        tm.sleep(0.5)
        # biblioteca time
except KeyboardInterrupt:
    print("\nCaptura interrompida")
    
# __________________________________________________SALVAR E PRINTAR ARQ ___________________________________________________________#      
  
wb.save(arq_excel)
print(arq_excel)
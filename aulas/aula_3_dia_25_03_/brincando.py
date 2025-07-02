#manupulação arquivos csv e excel

#CSV - tipo json - terá informação para criar novas infromações
#título título
#xxxxxx yyyyyy

#pandas - dataFrame - cria frames no cod
#openpyxl - biblioteca que permite editar e alterar, até mesmo cores

import pandas as pd
from openpyxl import Workbook, load_workbook   #workbooo é a tabela
from openpyxl.styles import PatternFill, Font



#criar dataFrame nome livros - lista com dict

livros = [
    {"Título": "Python", "Autor":"Arnaldo", "Ano": 2020},
    {"Título": "Python Cobra", "Autor":"Jorge", "Ano": 2025},
    {"Título": "Python Tython", "Autor":"Alma", "Ano": 2019},
]

#dataframe = df

df = pd.DataFrame(livros)
print("DataFrame criado:")
print(df)

#exportar para csv, com vírgula

df.to_csv("livros.csv", index=False)
print("Suuuucesssoooo")

df.to_excel("livros.xlsx", index=False) #não tem chave primária - se tiver deve por True
print("Suuuucesssoooo Excel")

#wb = workbook

#ler arquivo excel e manipular dados, para manipular deve criar dataframe

df_n = pd.read_excel("livros.xlsx")
df_n["Ano"] = df_n["Ano"]+1   #todo ano que pega ele soma 1 ex 1990 -> 1991
df_n.to_excel("Livro_atualizado.xlsx", index=False)
print("Atualizaou")

#formatar excel
#ws - worksheet - cria planilhas - ele cria sempre 3 planilhas

wb = load_workbook("Livro_atualizado.xlsx")
ws = wb.active

#estilo título

header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
header_font = Font(color="FFFFFF",bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

wb.save("Formatado.xlsx")
print("foi")
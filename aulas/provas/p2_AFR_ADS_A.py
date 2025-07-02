'''
Alessandra Furlanetto Rigonatti
RA: 2401151
ADS - Manhã - turma A

PROVA PRÁTICA DE RPA
'''
# pip install 

import requests
import sqlite3
import sys
import io

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

from bs4 import BeautifulSoup

from datetime import datetime

from reportlab.pdfgen import canvas

# ___________________ CARACTERES ESPECIAIS ____________________________________

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


# ___________________ EXTRAÇÃO DE DADOS DE 3 PAÍSES ____________________________________

def dadosDoPais():
    
    ''' Função  que irá fazer o crawber (rastrear e coletar dados, pelo que entendi) do país ecolhido pelo usuário'''
    
    local = input("🌎 Ecolhas o país que quer receber informações🧾: ")

    url = f"https://restcountries.com/v3.1/name/{local}"
    retorno =  requests.get(url)

    dados = retorno.json()

    dici = dados[0]
    nome = dici["name"]["common"]
    nomeOffi = dici["name"]["official"] 
    capital = dici["capital"][0] if "capital" in dici else "N/A"
    # se capital no dict mostre, se não capital marque como Not Available - Não disponível, não applicável
    continente = dici["continents"][0] if "continents" in dici else "N/A"
    regiao = dici["region"] if "region" in dici else "N/A"
    subRegiao = dici["subregion"]if "subregion" in dici else "N/A"
    populacao = dici["population"] if "population" in dici else "N/A"
    area = dici["area"] if "area" in dici else "N/A"
    fusoH = ", ". join(dici["timezones"]) if "timezones" in dici else "N/A" #aqui retorna lista, tive que converter para string
    bandeira = dici["flags"]["png"] if "flags" in dici and "png" in dici["flags"] else "N/A"

    # modeda é lista não dicionário então
    moeda = list(dici["currencies"].values())[0] if "currencies" in dici else {"symbol": "N/A", "name": "N/A"}
    simbo = moeda["symbol"]
    nomeMoed = moeda["name"]

    idioma = list(dici["languages"].values())[0] if "languages" in dici else "N/A"

    #apenas para mostar que está funconando
    print(f"Nome Oficial {nomeOffi}, Nome {nome} Capital {capital}\n Continente {continente} Região {regiao} Sub Região {subRegiao}\n Idioma {idioma} População {populacao}\n Área {area} Fuso Horário {fusoH}\n Moeda {simbo} {nomeMoed}\n Url bandeira {bandeira} ")

    # return dici
    # deve retornar lista ou tupla para o bd analisar
    return (nomeOffi, nome, capital, continente, regiao, subRegiao, idioma, populacao, area, fusoH, simbo, nomeMoed, bandeira)

# _________________________ BD
# sqlite3.connect -> conectar ou cria um bs, neste caso com nome bd_paises
# cursor -> executa comandos do sql, percorre os dados e retorna resultado
# execute -> executa tudo de uma vez, tipo docker-from django.conf import settings
# executemany -> executa muitas vezes vários registros
# ? -> PLACEHOLDER -> será subtituido pelo valor gerado
# commit() -> salva o bd
# close() -> sempre bom fechar a conexão com sql

def bd_paises():
    ''' Criação do banco de dados do retorno de paises'''
    
    conexao = sqlite3.connect('paises.db')
    resultado = conexao.cursor()
    
    resultado.execute('''
    CREATE TABLE IF NOT EXISTS paises (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nomeOffi  TEXT,
        nome TEXT,
        capital TEXT,
        continente TEXT,
        regiao TEXT,
        subRegiao TEXT,
        idioma TEXT,
        populacao INTEGER,
        area DOUBLE,
        fusoH TIMESTAMP,
        simbo TEXT,
        nomeMoed TEXT,
        bandeira TEXT)'''
        )
    
    for _ in range(3):
        info = dadosDoPais()  # "herdando" dados da função inicial
        resultado.execute('INSERT INTO paises (nomeOffi, nome, capital, continente, regiao, subRegiao, idioma, populacao, area, fusoH, simbo, nomeMoed, bandeira) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', info)   #deixar o mesmo numero de colunas, id não conta pois é automático
    
    conexao.commit()
    conexao.close()


# ___________________ EXTRAÇÃO DE DADOS DE 10 LIVROS ____________________________________
# parser = analisar -> html.parser analisa o html


def dez_titulos():
    '''Percorrer os 10 primerios titulos da 1° pag e extrair infromações sobre o livro'''
    
    try:
        url = "https://books.toscrape.com/catalogue/page-1.html"
        retorno = requests.get(url)
        
        if retorno.status_code != 200:
            return []
            
        soup = BeautifulSoup(retorno.text, "html.parser")
        livros = []
        
        livros_listados = soup.find_all("article", class_="product_pod")[:10]  # direto no for estava dando erro
        
        for dados in livros_listados:  # percorrerá penas os 10 primeiros
            titulo = dados.h3.a["title"]
            preco = dados.find("p", class_="price_color").get_text()
            estoque = "icon-ok" if dados.find("p", class_="instock availability") else "Out of stock"
            aval_dado = dados.find("p", class_="star-rating")
            aval_estre = {"one": "⭐","Two": "⭐⭐", "Three": "⭐⭐⭐", "Four": "⭐⭐⭐⭐", "Five": "⭐⭐⭐⭐⭐" }
            aval = aval_estre.get(aval_dado["class"][1])
            
            livros.append(f" titulo {titulo}, preço {preco}, estoque {estoque} e valiação {aval}")
            
            #dict parar acesso
            livros.append({
                "titulo": titulo,
                "preco": preco,
                "estoque": estoque,
                "aval": aval
            })
            
        return livros
    
    except Exception as e:
        print(f"Erro na extração: {e}")
        return []

# _________________________ BD

def bd_livros():
    ''' Criação e salvamento do bd de livros'''
    
    conexão = sqlite3.connect('titulos_livros.db')
    resultado = conexão.cursor()
    
    resultado.execute('''
    CREATE TABLE IF NOT EXISTS livros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT,
        preco REAL,
        estoque TEXT,
        aval TEXT)'''
        )
    
    info = dez_titulos()
    for livro in info:
        try:
            resultado.execute('''INSERT INTO livros (titulo, preco, estoque, aval) VALUES (?, ?, ?, ?)''', (livro["titulo"], livro["preco"], livro["estoque"], livro["aval"]))
        except Exception as e:
            print(f"Erro ao inserir livro: {e}")  #aval e não acal
    
    conexão.commit()
    print("dados inseridos")
    
    conexão.close()


# ___________________ GERAR RELATÓRIO EXCEL ____________________________________
# df = data Frame (praticamente uma tabela com colunas e linhas)
# wb = Work Book/ Pasta de Trabalho - amazena objeto na planilha
# ws = Worksheet/ Planilha - armazena planilha ativa  na pasta de trabalho


def relatorioExcel():
    '''junção dos bancos de dados criados em outras funções e geração do relatório'''
    
    wb = Workbook()
    data_atual = datetime.now().strftime("%d-%m-%Y")
    
#_____________ PAISES
    con = sqlite3.connect("paises.db")
    df_paises = pd.read_sql("SELECT * FROM paises", con)
    con.close()
    
    ws_paises = wb.active
    ws_paises.title = "Países"
    
    for r in dataframe_to_rows(df_paises, index=False, header=True):
        ws_paises.append(r)
    
#________________ LIVROS   
    con = sqlite3.connect("titulos_livros.db")
    df_livros = pd.read_sql("SELECT * FROM livros", con)
    con.close()
    
    ws_livros = wb.create_sheet("Livros")
    
    for r in dataframe_to_rows(df_livros, index=False, header=True):
        ws_livros.append(r)
    
#________________ DADOS E CABEÇALHO

# sheet = folha    create.sheet é criar uma folha, no caso planilha
# lembrar que deve fazer for sheet porque percorre celulas e não linhas
# para não esquecer - [""]*3 = 3 celular vazias

    data_atual = datetime.now().strftime("%d-%m-%Y")
    ws_info = wb.create_sheet(f"Relatório Final {data_atual}")


    ws_info.append(["Relatório Excel P2 - RPA - ADS - A - Manhã", "", "", ""])  
    ws_info.append(["Aluno e RA:", "Alessandra - 2401151", "", ""])              
    ws_info.append(["Data da Criação:", datetime.now().strftime("%d/%m/%Y"), "", ""])  

    #__________ MESCLAR E CENTRALIZAR HEADER, LINHAS 1 E 2

    ws_info.merge_cells('A1:D1')  
    ws_info['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_info.merge_cells('B2:D2')  
    ws_info['B2'].alignment = Alignment(horizontal='left')  # Alinha - esquerda 
    ws_info.merge_cells('B3:D3')  # Mescla B3 até D3
    ws_info['B3'].alignment = Alignment(horizontal='left')  

    for row in ws_info.iter_rows(min_row=1, max_row=3, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
        
    #____________ SALVAR
        data= datetime.now().strftime("%d-%m-%Y")
        wb.save(f"relatorio_AFR_ADS_A {data}.xlsx")
        print("Relatório gerado com sucesso!")

#___________ CONVERSÃO
def dataframe_to_rows(df, index=False, header=True):
    if header:
        yield list(df.columns)
    for _, row in df.iterrows():
        yield list(row)

#__________________________________FUNÇÃO MAIN__________________________________

def main():
    bd_paises()
    bd_livros()
    relatorioExcel()

#____________________________________ MAIN____________________________________

if __name__ == "__main__":
    main()

#_______________________________ OBS _____________________________________________

# https://books.toscrape.com/catalogue/
# a-light-in-the-attic_1000/index.html
# clicar f12 /inspecionar
# parte superiora esquerda clicar em SELECT AN ELEMENT (QUADRADO PONTILHADO COM UMA SETA OU CTRL SHIFT C)
# Na pag que está o nome do livro, o valor, passar mouse em cima, verá que formará um caixa clorida. Em elements, irá mostrar a linha que está o dado
# ex: <p class="price_color">E51.77</p> -> usará price = soup.find('p', class_='price_color').text para extrair o valor
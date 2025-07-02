import pyautogui as pag
# biblioteca .py - interface gráfica - controle de mouse, teclado e interação com atela
import time as tm
# biblioteca de manipulação de tempo, ex sleep
import os
# função que interage com so (sistema operacional) - gerenciamento de files como criar, deletar e seus caminhos
import pandas as pd
# biblioteca para nanipular dados
from openpyxl import load_workbook
# carrega planilha excel .xlsx
# worklbook = tabela
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
# Font fontes e PatternFill cores de fundos
from openpyxl.utils import get_column_letter
# tem haver com colunas
import tkinter as tk 
# biblioteca padrão do py que criar interfaces gráficas (GUIs - Graphical User Interface - Interface Gráfica do Usuário) ex botões, menus e aumentar emojis 



# ___________________________________ CRIANDO XLSX ___________________________________________________________________#

def tabela_base():
    tabela = {
        "tarefa":["Abrir file", "salvar file", "fechar file", "editar file"],
        "tipo":["click", "hotkey", "hotkey", "manual"],
        "dados":["wind+bloco de notas", "ctrl+s", "alt+f4", "---"]
    }
    
    return tabela

#cuidado com pontuação errada : iníco na chave e , entre dados
#editar será feito pelo usuário

# _____________________________________________ CRIANDO DF/DATAFRAME ____________________________________________________________#

def dataFrame ():
    
    df = pd.DataFrame(tabela_base())
    return df                            


#dentro do objeto dataframe passar as chaves/ tarefa, tipo e dados, conteúdo da coluna é o que está dentro das chaves
# ATENÇÃO CHAMA A FUNÇÃO COLOCANDO NOME E ()

# _____________________________________________ SALVANDDO EM EXCEL _______________________________________________________________#

def salvar_arq():
    arq_excel = "tabela.xlsx"
    df.to_excel(arq_excel, index=False, sheet_name="tabela_RPA")
    return arq_excel

# df.to_excel = transformando em arq excel | sheet_name = 1° planilha se chamará tabelaRPA (mudei o nome do doc salvo)
# arq_excel é minga tabela.xlsx | index= False = não tem chave primária - se tiver deve por True

# __________________________________________________ FORMATAR ARQ _________________________________________________________________#

def formatar_tab(arq_excel):
    wb = load_workbook(arq_excel)
    ws = wb["tabela_RPA"]

    # wb = workbook
    # load_workbook = carrega o arq excel
    # ws = worksheet - aba específica - lembrando que ao criar vem com 3 abas


    # variável header_fill = PatternFill(start_color="edede9", end_color="f5ebe0", fill_type="gradient") # fundo de gradiente
    #cor row = linha completa
    header_font = Font(color="253237",bold=True)
    header_fill = PatternFill(start_color="d8e4e8", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    # configuração e aplicação - cabeçalho
    # usa-se o for cel porque deverá, na 1° linha, percorrer todas suas células e aplicar, "for" horizontal
    #cor row = linha completa

    #linha zebrada com sim cor não
    lin_par_c = PatternFill(start_color="e0fbfc", fill_type="solid")
    lin_inp_c = PatternFill(start_color="c2dfe3", fill_type="solid") 
    for row in ws.iter_rows(min_row=2): 
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = lin_par_c
            else:
                cell.fill = lin_inp_c


    # header_fill/ define cor de fundo - start color = inicia com esta cor   | fill_type = forma/estilo de preenchimento da cálula  
    # end_color = termina a cor | quando usa start e end juntos dá degradê, o que significa que fill_type é tipo gradinete (algo assim)
    # header_fonte = fonte do cabeçalho, pode ser color, bold, italic, size e name(qual fonte)
 
    for coluna in ws.columns:
            tamanho_max = 0
            for celula in coluna:
                if len(str(celula.value)) > tamanho_max:
                    tamanho_max = len(str(celula.value))
            ws.column_dimensions[coluna[0].column_letter].width = tamanho_max + 2
        
            wb.save(arq_excel)
    #retorna todas colunas do ws/worksheet ex células, em cada coluna
    # inicia com lagurga 0 na coluna
    # para cada céluna na coluna
    # str(celula.value) → Converte o valor para string - converte para string
    # len(str(celula.value)) = percorre/calcula tamanho do texto e vê se é maior que tamanho máximo do texto
    # se o tamanho for maior a coluna irá aumentar - ex celular da linha 7 da coluna 2 , o valor da len é 5, então vai atualizar para 5 a largura
    # ws.column_dimensions[...].width → Define a largura da coluna.- largura da coluna
    # manho_max + 2 → Adiciona 2 estaços para não ficar rente na linha
    
    
    #wb.save salva aterações no formato do arq existente na função
    #como é função de "ação" mão se usa return, apenas os "comandos/cods"

# _____________________________________________________ POSIÇÃO DO MOUSE ___________________________________________________________#  
def mouse ():
    try:
        while True:
            x, y = pag.position()
            # biblioteca pyautogui = pug  
            # print(f"\nLocal atual x = {x} e y = {y}", end="\r") -> aqui dá uma lista de posições, quero apenas posição atual
            print(f"Local atual mouse x = {x} e y = {y} | Tecle no terminal e ctrl+C para sair", end="\r")
            #aqui muda a posição sem mostrar o "histórico" de posição e comando para sair
            #dependendo do posição aparece sair, sairr, sairrr, pesquisar erro
            tm.sleep(0.5)
            # biblioteca time
    except KeyboardInterrupt:
        print("\nCaptura interrompida")

# __________________________________________________SALVAR E PRINTAR ARQ ___________________________________________________________#      
def salvar_exc(arq_excel): 
    wb.save(arq_excel)  
    print(arq_excel)   

# ___________________________________________________ AUTOMATIZANDO _________________________________________________________________#
print("Automatizando projeto bloco notas")


# __________ ABRIR BLOCO __________

def abrir_bloco(zerado = False):
    
    if zerado:
        pag.press("win")
        pag.write("bloco de notas")
        pag.press("enter")
        
    else:
        if df.iloc[0]["tipo"] == "click":
            pag.press("win")
            pag.write("bloco de notas")
            pag.press("enter")
            
    tm.sleep(2)
    
#iloc = Integrer Location - localiza por int -  acessa linhas do dataframe (df), neste caso [0] é a linha zero 1°, sem ser cabeçalho
#seria | 0 |abrir file | click | wind+bloco de notas |
#se dataframe na localização linha 0, coluna tipo for igual a "click"
#automaticamente pressione tecla win, depois escreva bloco de notas, sem seguida tecle enter e espera 2 segundo para próximo comando


# __________ USUÁRIO DIGITANDO __________

def digitar():
    print("\nTem 7 segundos para digitar seu texto ")
    tm.sleep(7)

#de novo, aqui é cód de ação, sem retorno

# __________ SALVANDO __________

def salvar():
    if df.iloc[1]["tipo"] == "hotkey":
        pag.hotkey(*df.iloc[1]["dados"].split("+"))
        tm.sleep(1)
        pag.write("Bloco Automatizado RPA")
        pag.press("enter")
        print(f" Arquivo Bloco Automatizado RPA foi salvo com sucesso ")

# se o dataframe. Integrer Location/ localização por int na linha 1 (1 | salvar file | hotkey | ctrl+S |) na coluna tipo for hotkey(acho que é #chave especial) : pyautogui.hotkey irá desempacotar na linha um na coluna dados o ctrl S(separa ctrl, s) | depois de 1segundo | escreva o mome
# do arquivo (neste caso bloco Automático RPA) (writer)| ao pressionar (press) enter tudo será salvo e mostrará mesangem de salvo

#iloc = Integrer Location - localiza por int -  acessa linhas do dataframe (df), neste caso [0] é a linha zero 1°, sem ser cabeçalho
#seria | 0 |abrir file | click | wind+bloco de notas |
# * = desenpacota a lista em argumentos separados, isolados | WIND SEPARADO DE BLOCO DE NOTAS
# split - divide string em partes  -> split(+) -> entendi que win+cloco de notas -> + seria o caracteres especial que mostra onde será
# a divisão das string -> na hora rodas será pressionado o win depois bloco de notas
# pag = pyautogui | hotkey seria junção de teclas ex crtl C, neste caso tecla windows e bloco(não é bem uma tecla)
#tm = time | sleep tempo de espera para executar próxima função | (2) 2 seg

# __________ FECHANDO __________

def fechar():
    if df.iloc[2]["tipo"] == "hotkey":
        tm.sleep(1)
        pag.hotkey(*df.iloc[2]["dados"].split("+"))

print("Automatização feita - tabela + doc excel criado e tabela usada para abrir e editar bloco de notas")


# _____________________________________________ RELATÓRIO ________________________________________________________________________#

def gerar_relatorio():
    
    
# __________ CRIANDO TABELA __________    
    wb = Workbook()
    planilha = wb.active
    planilha.title = "Relatório"
#wb recebe workbook = todo o arquivo
#variável planilha recebe wb - planilha ativa/ visível
#variável planilha_titulo recebe nome relatório (do arquivo)
    
# __________ FORMAT CABEÇ, FONT __________
    cabecalho = ["Tarefa", "Status", "Tempo (s)"]
    planilha.append(cabecalho)
# adicionando cabeçalho a planilha    

    for celula in planilha[1]:
        celula.font = Font(bold=True)
# percorra as células na planilha, inicinado na linha 1 
# deixe as fonte em bold (neste caso a cor será automaticamente preta)
   
# __________ DADOS __________
    tarefas = [
        {"Tarefa": "Abrir Bloco de Notas", "Status": "Concluído", "Tempo (s)": 2},
        {"Tarefa": "Tempo de Digitação", "Status": "Concluído", "Tempo (s)": 7},
        {"Tarefa": "Salvar Arquivo", "Status": "Concluído", "Tempo (s)": 1},
        {"Tarefa": "Fechar Bloco de Notas", "Status": "Concluído", "Tempo (s)": 1}
    ]
 # criei dict com lista de tarefas e o tempo delas
 
    for tarefa in tarefas:
        planilha.append([tarefa["Tarefa"], tarefa["Status"], tarefa["Tempo (s)"]])
# para cada tarefa no dict tarefas adicione tarefa, satus e tempo em segundo
        
# __________ FORMAT LARG __________
    for coluna in planilha.columns:
        tamanho_max = 0
        for celula in coluna:
            if len(str(celula.value)) > tamanho_max:
                tamanho_max = len(str(celula.value))
        planilha.column_dimensions[coluna[0].column_letter].width = tamanho_max + 2
#retorna todas colunas do ws/worksheet ex células, em cada coluna
    # inicia com lagurga 0 na coluna
    # para cada céluna na coluna
    # str(celula.value) → Converte o valor para string - converte para string
    # len(str(celula.value)) = percorre/calcula tamanho do texto e vê se é maior que tamanho máximo do texto
    # se o tamanho for maior a coluna irá aumentar - ex celular da linha 7 da coluna 2 , o valor da len é 5, então vai atualizar para 5 a largura
    # ws.column_dimensions[...].width → Define a largura da coluna.- largura da coluna
    # manho_max + 2 → Adiciona 2 estaços para não ficar rente na linha
        
# __________ SALVANDO __________
    nome_arquivo = "Relatorio.xlsx"
    wb.save(nome_arquivo)
    return nome_arquivo

    
    
###############################################################################################################################

# __________________________________________ MAIN ____________________________________________________________________________#

def main():
    global df 
    
    print("Bem vindo/s a primeira automatização de Alessandra. Espero que goste/m")
    
   
    df = dataFrame()
    
    
    arq_excel = salvar_arq()
    formatar_tab(arq_excel)
    
    abrir_bloco(zerado=True)
   
    print("\nBloco de notas ativado")
    abrir_bloco()
    digitar()
    salvar()
    fechar()
    
    relatorio = gerar_relatorio()
    print(f"\n📊 Relatório gerado em: {os.path.abspath(relatorio)}")     
    print("     🤖       Automação completa      💾") 
    

if __name__ == "__main__":
    main()
    
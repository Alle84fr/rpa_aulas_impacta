import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 1. Dados fictícios de presença
dados = {
    'Aluno': ['André', 'Bianca', 'Caio', 'Daniela'],
    'Aulas Assistidas': [18, 15, 12, 20],
    'Total de Aulas': [20, 20, 20, 20],
}
df = pd.DataFrame(dados)

# 2. Calcular % de presença no formato decimal (0.75 para 75%)
df['% Presença'] = df['Aulas Assistidas'] / df['Total de Aulas']

# 3. Salvar em Excel
arquivo = 'controle_presenca.xlsx'
df.to_excel(arquivo, index=False, sheet_name='Presenças')

# 4. Formatar com openpyxl
wb = load_workbook(arquivo)
ws = wb['Presenças']

# 5. Estilo do cabeçalho (cinza claro + negrito)
header_fill = PatternFill(start_color='DDDDDD', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# 6. Formatar % como percentual (duas casas)
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):  # Coluna '% Presença'
    for cell in row:
        cell.number_format = '0.00%'

# 7. Destacar alunos com presença abaixo de 75% (vermelho claro)
for row in ws.iter_rows(min_row=2, max_col=4):
    presenca = row[-1].value
    if presenca is not None and presenca < 0.75:
        row[-1].fill = PatternFill(start_color='FF9999', fill_type='solid')

# 8. Autoajustar largura das colunas
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2

# 9. Salvar alterações
wb.save(arquivo)
print(f"Arquivo '{arquivo}' salvo com sucesso.")
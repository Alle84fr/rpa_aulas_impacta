import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 1. Criar DataFrame com dados fictícios
dados = {
    'Nome': ['Ana', 'Bruno', 'Carlos', 'Diana'],
    'Nota 1': [8.5, 7.0, 6.0, 9.0],
    'Nota 2': [9.0, 6.5, 7.5, 8.0],
}
df = pd.DataFrame(dados)

# 2. Calcular média com pandas
df['Média'] = df[['Nota 1', 'Nota 2']].mean(axis=1)

# 3. Salvar em Excel com pandas
arquivo_excel = 'notas.xlsx'
df.to_excel(arquivo_excel, index=False, sheet_name='Notas')

# 4. Abrir com openpyxl para formatar
wb = load_workbook(arquivo_excel)
ws = wb['Notas']

# 5. Formatar cabeçalho: negrito + preenchimento cinza
header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# 6. Formatar colunas numéricas para 2 casas decimais
for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '0.00'

# 7. Destacar notas abaixo de 7 com fundo vermelho claro
for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value < 7:
            cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")

# 8. Autoajustar largura das colunas
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_length + 2

# 9. Salvar alterações
wb.save(arquivo_excel)

print(f"Planilha '{arquivo_excel}' criada e formatada com sucesso!")
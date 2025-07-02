import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 1. Dados de vendas
dados = {
    'Vendedor': ['Alice', 'Bob', 'Carol', 'David'],
    'Janeiro': [1200, 850, 970, 1300],
    'Fevereiro': [1500, 900, 1100, 1250],
    'Março': [1600, 1000, 1050, 1400]
}
df = pd.DataFrame(dados)

# 2. Total por vendedor
df['Total'] = df[['Janeiro', 'Fevereiro', 'Março']].sum(axis=1)

# 3. Salvar Excel
arquivo = 'relatorio_vendas.xlsx'
df.to_excel(arquivo, index=False, sheet_name='Vendas')

# 4. Formatar com openpyxl
wb = load_workbook(arquivo)
ws = wb['Vendas']

# Cabeçalho negrito + preenchimento azul claro
header_fill = PatternFill(start_color='B7D8F1', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# Formatar valores como moeda
for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
    for cell in row:
        cell.number_format = 'R$ #,##0.00'

# Destaque vendedores com total acima de R$ 4000
for row in ws.iter_rows(min_row=2, max_col=6):
    total_cell = row[-1]
    if total_cell.value is not None and total_cell.value > 4000:
        total_cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')  # Verde claro

# Autoajustar colunas
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_len + 2

wb.save(arquivo)
print(f"Arquivo '{arquivo}' criado com sucesso.")
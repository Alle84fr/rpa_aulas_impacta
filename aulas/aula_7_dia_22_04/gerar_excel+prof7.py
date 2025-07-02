# pandas, openpyxl são as biblioteca de geração do Excel
# pip install nome_Biblioteca

#Importar as bibliotecas
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 1- criar o DataFrame e salvar o arquivo
dados = {
    'Produto': ['Caneta', 'Lápis', 'Caderno', 'Borracha'],
    'Preço' : [2.50, 1.20, 15.00, 0.80]
}

df = pd.DataFrame(dados)

#Salvar o arquivo Excel
df.to_excel('produtos.xlsx', index=False, engine='openpyxl')

#2 - Reabrir o arquivo Excel com openpyxl e aplicar a formatação
wb = load_workbook('produtos.xlsx') 
ws = wb.active #seleciona a planilha ativa

#3- Deixar a primeira linha em negrito
for cell in ws[1]: # ws[1] acessa a primeira linha inteira
    cell.font = Font(bold=True)

wb.save('produtos_formatado.xlsx')
print("Arquivo criado com sucesso!")